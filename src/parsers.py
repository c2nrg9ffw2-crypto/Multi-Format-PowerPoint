"""Parse .xlsx and .pdf source files into structured dicts."""
import sys
from pathlib import Path


class ParseError(ValueError):
    """Raised when a source file cannot be read or is unsupported."""


def parse_xlsx(path: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        raise ParseError("openpyxl is not installed. Run: pip install openpyxl")

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise ParseError(f"Cannot open {path!r}: {e}") from e

    result = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append([str(c) if c is not None else "" for c in row])
        if rows:
            result[sheet] = rows[:501]  # header + 500 data rows

    if not result:
        print(f"Warning: {path} contains no data.", file=sys.stderr)

    return result


def _extract_tables(page) -> list[list[list[str]]]:
    tables = []
    for table in page.extract_tables():
        cleaned = [
            [str(cell).strip() if cell else "" for cell in row]
            for row in table
            if any(cell for cell in row)
        ]
        if cleaned:
            tables.append(cleaned)
    return tables


def parse_pdf(path: str) -> dict:
    try:
        import pdfplumber
    except ImportError:
        raise ParseError("pdfplumber is not installed. Run: pip install pdfplumber")

    try:
        pages = []
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = (page.extract_text() or "").strip()
                tables = _extract_tables(page)
                pages.append({"index": i, "text": text, "tables": tables})
    except ParseError:
        raise
    except Exception as e:
        raise ParseError(f"Cannot open {path!r}: {e}") from e

    if not pages:
        print(f"Warning: {path} contains no pages.", file=sys.stderr)

    return {"pages": pages}
